# 3단계
# 1 ~ N 페이지까지 크롤링
import time

import requests
from bs4 import BeautifulSoup
from icecream import ic
import pandas as pd
# from openpyxl.reader.excel import load_workbook
from openpyxl import load_workbook

# input_page = 3

# input_page = int(input("몇 페이지까지 수집하시겠습니까? >>> "))
input_text = input("몇 페이지까지 수집하시겠습니까? >>> ")
if input_text == "":
    input_page = 1
else:
    input_page = int(input_text)

data = []
# for page in range(1, input_page):
for page in range(1, input_page + 1):
    response = requests.get(f"https://kin.naver.com/search/list.naver?query=%EC%82%BC%EC%84%B1%EC%A0%84%EC%9E%90&page={page}")
    html = response.text
    soup = BeautifulSoup(html, 'html.parser')

    # 나무 태그 찾기
    articles = soup.select(".basic1 > li")

    i = 1
    for article in articles:
        ic(i * page)

        title = article.select_one("._nclicks\\:kin\\.txt").text.strip()
        ic.disable()
        ic(title)

        title = article.select_one("._searchListTitleAnchor").text.strip()
        ic(title)

        title_select = article.select_one("._nclicks\\:kin\\.txt._searchListTitleAnchor")
        title = title_select.text.strip()
        ic(title)

        link = title_select.attrs['href']
        ic(link)

        date = article.select_one(".txt_inline").text
        ic(date)

        category = article.select_one(".txt_block > a:nth-of-type(2)").text
        ic(category)

        # reply_count = soup.select_one(".txt_block > span:nth-of-type(2)").text.split(' ')[1]
        reply_count = article.select_one(".txt_block > span:nth-of-type(2)").text.split("답변수")[1].strip()
        ic(reply_count)

        ic.enable()
        ic(title, link, date, category, reply_count)

        data.append([title, link, date, category, int(reply_count)])

        i += 1

    time.sleep(1)

# 데이터 프레임 생성
df = pd.DataFrame(data, columns=['제목', '링크', '날짜', '카테고리', '답변수'])

# 엑셀 저장
excel_path = "naver_finance_crawling.xlsx"
# df.to_excel(excel_path, engine="openpyxl", index=False, freeze_panes=(1, 1))
df.to_excel(excel_path, engine="openpyxl", index=False)

# 엑셀 파일 열기
wb = load_workbook(excel_path)
ws = wb.active

# 열 너비 설정
ws.column_dimensions['a'].width = 70
ws.column_dimensions['b'].width = 100
ws.column_dimensions['c'].width = 30
ws.column_dimensions['d'].width = 30
ws.column_dimensions['e'].width = 10

# 각 열의 너비를 200으로 설정
# for column_cells in ws.columns:
#     length = 50
#     col_letter = column_cells[0].column_letter
#     ws.column_dimensions[col_letter].width = length

# 엑셀 파일 저장
wb.save(excel_path)

# 한글 깨짐 방지, index 번호 컬럼 미노출
# df.to_csv("naver_kin_crawling.csv", index=False, encoding="utf-8-sig")